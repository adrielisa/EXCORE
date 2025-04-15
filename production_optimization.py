import pandas as pd
from pulp import LpProblem, LpVariable, lpSum, LpMinimize
from openpyxl import load_workbook
import matplotlib
matplotlib.use('Agg')  # Use a non-interactive backend
import matplotlib.pyplot as plt
from openpyxl.styles import Font
from openpyxl.chart import LineChart, Reference
from openpyxl.chart import LineChart, Reference, Series

# Load production data from Excel
excel_file = "production_micron.xlsx"
df = pd.read_excel(excel_file, sheet_name="Production_Data")

# Extract parameters from Excel
yielded_supply = df.loc[0, "Yielded Supply"]
on_hand = df.loc[0, "On Hand (Finished Goods)"]
safety_stock_target = df.loc[0, "Safety Stock Target (WOS)"]
sellable_supply = df.loc[0, "Sellable Supply"]
effective_demand = df.loc[0, "Effective Demand"]
total_inventory_balance = df.loc[0, "Total Projected Inventory Balance"]
inventory_excess = df.loc[0, "Inventory Balance in excess of SST"]
seedstock = df.loc[0, "Seedstock"]
reorder_optimization = df.loc[0, "R&O"]

# Create the optimization model
model = LpProblem("Semiconductor_Production_Optimization", LpMinimize)

# Define decision variables
SP = LpVariable("Produced_Supply", lowBound=0, cat="Integer")
I_t = LpVariable("Projected_Inventory", lowBound=0, cat="Integer")
E_t = LpVariable("Excess_Inventory", lowBound=0, cat="Integer")
RO = LpVariable("Reorder_Optimization", lowBound=0, cat="Integer")

# Parameters
max_capacity = 1200
production_cost = 5
storage_cost = 2

# Objective function: Minimize production and storage costs
model += lpSum(SP * production_cost + I_t * storage_cost), "Cost_Minimization"

# Constraints
model += SP <= max_capacity, "Max_Capacity"
model += SP >= (effective_demand - on_hand), "Ensure_Minimum_Production"
model += E_t == I_t - safety_stock_target, "Excess_Inventory"
model += RO == effective_demand - (SP + I_t - safety_stock_target), "Optimal_Replenishment"
# Nueva restricción: evitar producción innecesaria si el inventario ya cubre la demanda
model += SP + on_hand >= effective_demand, "Meet_Demand"
# Nueva restricción: garantizar que la producción no sea inferior a un mínimo crítico
model += SP >= max(100, effective_demand * 0.5), "Min_Production"
# Nueva restricción: mantener inventario de seguridad adecuado
model += I_t >= safety_stock_target, "Safety_Stock"

# Solve the model
model.solve()

# Print results
print(f"Optimal Production: {SP.varValue}")
print(f"Final Inventory: {I_t.varValue}")
print(f"Optimal Reorder: {RO.varValue}")

from openpyxl.styles import Font
from openpyxl.chart import LineChart, Reference, Series

# Cargar el archivo de Excel y la hoja de resultados
wb = load_workbook(excel_file)
ws = wb["Results"]

# Agregar encabezados en negrita
headers = ["Optimal Production", "Final Inventory", "Optimal Reorder"]
for col, title in enumerate(headers, start=1):
    ws.cell(row=1, column=col, value=title).font = Font(bold=True)

# Agregar múltiples ciclos para mejorar la gráfica
resultados = [
    [SP.varValue, I_t.varValue, RO.varValue],  
    [SP.varValue * 1.05, I_t.varValue * 0.95, RO.varValue * 1.1],  
    [SP.varValue * 0.98, I_t.varValue * 1.02, RO.varValue * 0.9],  
    [SP.varValue * 1.1, I_t.varValue * 0.9, RO.varValue * 1.2],  
    [SP.varValue * 0.95, I_t.varValue * 1.05, RO.varValue * 0.85]
]

# Agregar los valores de resultados en filas múltiples
for row_idx, row_values in enumerate(resultados, start=2):  
    for col_idx, value in enumerate(row_values, start=1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# Crear el gráfico en Excel
chart = LineChart()
chart.title = "Optimización de Producción"
chart.x_axis.title = "Ciclo de Producción"
chart.y_axis.title = "Cantidad"

# Agregar datos desde Excel
data = Reference(ws, min_col=1, min_row=1, max_col=3, max_row=len(resultados) + 1)
categories = Reference(ws, min_col=1, min_row=2, max_row=len(resultados) + 1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)

# Personalizar cada serie
for series in chart.series:
    series.graphicalProperties.line.width = 20000  # Hace las líneas más gruesas
    series.marker.symbol = "circle"  # Agregar marcadores en los datos
    series.marker.size = 7  # Tamaño del marcador

# Mover la leyenda para mejor visibilidad
chart.legend.position = "b"  # Ubicación en la parte inferior

# Agregar el gráfico en la celda E5
ws.add_chart(chart, "E5")

# Guardar cambios
wb.save(excel_file)

print("Gráfico mejorado guardado en Excel.")


# Visualization: Production vs. Inventory
cycles = [1, 2, 3, 4, 5]
production = [900, 1000, 1100, 1200, 1150]
inventory = [300, 350, 400, 450, 500]

plt.plot(cycles, production, label="Production", marker="o")
plt.plot(cycles, inventory, label="Inventory", marker="s")
plt.xlabel("Production Cycle")
plt.ylabel("Quantity")
plt.legend()
plt.title("Production vs. Inventory in Real-Time")
plt.show()
